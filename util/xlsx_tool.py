import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.styles import Font

def xlsx_ws_bold_row(ws, row):
    for cell in ws[row]:
        cell.font = Font(bold=True)

def xlsx_ws_bold_col(ws, col:str):
    for cell in ws[col]:
        cell.font = Font(bold=True)

def xlsx_ws_neg_bg(ws, min_row, min_col):
    # Set fill color for negative numbers
    redFill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    for row in ws.iter_rows(min_row=min_row, min_col=min_col):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.fill = redFill

def xlsx_col_fit(wb):
    for ws in wb:
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name
            # Since Openpyxl 2.6, the column name is  ".column_letter" as .column became the column number (1-based)
            for cell in col:
                try:  # Necessary to avoid error on empty cells
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 5) * 1.05
            ws.column_dimensions[column].width = adjusted_width
