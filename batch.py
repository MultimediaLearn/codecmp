import os
import csv
import glob

import score
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

from util.vutil import *
from arguments import *
from util.xlsx_tool import *

def save_refs(res, ws, ws_fig):
    # print(res)
    head = ["class", "file", "psnr_avg(FF)", "psnr_y(VMAF)", "ssim_all(FF)", "ssim(VMAF)", "vmaf(VMAF)",
            "time_save(u+s)%", "enc_name", "label", "par"]
    if ws.dimensions == "A1:A1":
        print(head)
        ws.append(head)
        xlsx_ws_bold_row(ws, 1)
    fig_pos_row = 1
    for ref_key in res:
        ref_res = res[ref_key]
        bd_refs = ref_res.bdres
        _, yuv_file = os.path.split(ref_key)
        for par_key in bd_refs:
            enc_name, bd, bd_fig = bd_refs[par_key]
            content = [ ref_res.fclass, yuv_file,
                round(bd["psnr_avg"], 5), round(bd["psnr_y"], 5),
                round(bd["ssim_all"], 5), round(bd["ssim"], 5), round(bd["vmaf"], 5),
                round(bd["time"] * 100, 2), enc_name, ref_res.flabel, par_key
                ]
            ws.append(content)
            print(content)
            anchor_cell = "A" + str(fig_pos_row)
            ws_fig[anchor_cell] = ref_res.fclass
            ws_fig[anchor_cell].font = Font(bold=True)
            anchor_cell = "B" + str(fig_pos_row)
            ws_fig[anchor_cell] = yuv_file
            ws_fig[anchor_cell].font = Font(bold=True)
            anchor_cell = "A" + str(fig_pos_row + 1)
            img = openpyxl.drawing.image.Image(bd_fig)
            ws_fig.add_image(img, anchor_cell)
            fig_pos_row += 20
    xlsx_ws_neg_bg(ws, 1, 1)
    xlsx_ws_bold_col(ws, "A")

if __name__ == "__main__":
    print("Input arguments list:")
    print(pretty_args(args, tabs="  "))

    enc = args.enc
    refs = args.refs
    resume = args.resume
    out_dir = "out_" + str(args.id) + "/"
    res_file = os.path.join(out_dir, args.res)

    init_logger(level=logging.INFO, logfile=res_file + "." + str(args.is_init) + ".log")
    pwarn("Input arguments list:")
    pwarn(pretty_args(args, tabs="  "))

    wb = Workbook()
    ws_bdrate = wb.active
    ws_bdrate.title = "bdrate"
    ws_fig = wb.create_sheet("figure")
    ws_infos = wb.create_sheet("details")

    # create path directory
    [res_path, _, _] = sep_path_segs(res_file)
    if (res_path):
        make_dir(res_path)

    res = score.eval(enc, refs, resume, wb)
    save_refs(res, ws_bdrate, ws_fig)
    xlsx_col_fit(wb)
    wb.save(res_file)
